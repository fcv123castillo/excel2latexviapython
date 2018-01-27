# EXCEL 2 LATEX SUB-FUNCTIONS
########################################################################################################################
# This library file contains all the functions used by the Excel2LaTeXviaPython code

import re           # For reading and processing text strings
import openpyxl     # For obtaining the Excel file formatting


def is_number(s):
    """
    IS_NUMBER

    This function tells us if the object "s" is a number or not.

    This is useful for working out if the excel cell contains a number, and hence whether we need to check how many d.p.
    to round to. It checks by trying to convert the string to a float. If it fails, "s" is not a number

    Args:
        s: [string]

    Returns:
        True/False: Answers if "s" is a string containing only a number
    """

    try:
        float(s)
        return True
    except ValueError:
        return False


def cell_is_value(s, search=re.compile(r'[^0-9\*\.()+-eE]').search):
    """
    We only want to round numbers in the table if it is comes from a cell that is only a number, and not from any 
    potential table headers cells than contain a number. So check to see if the cell contains only numbers, decimal
    points, parenthesis, asterisk. In which case, return yes. These other terms are included as often tables report
    standard errors in parenthesis or include asterisk to denote statistical significance.

    I added "e" and "E" into the list of characters to allow for scientific notation. Currently the code rounds off
    scientific notation to the specified number of DPs. Should probably create a setting to allow for scientific
    notation to be preserved.
    
    :param s: string of a number that is to be rounded
    :return: True/False: Answers if "s" is from a cell primarily a number
    """

    return not bool(search(s))


def clean_cell_str(s):
    """
    CLEAN_CELL_STR

    Sometimes (especially if the excel spreadsheet is generated by another program), cells contents are actually
    formatted as [="CONTENTS"]. If this is the case, we want to return only the CONTENTS and not the =" at the start
    or " at the end.

    Args:
        s: [string]

    Returns:
        string containing only the true content of the cell
    """

    if re.search('=".*"', s):
        return s[2:-1]
    else:
        return s


def tupple2latexstring(row_tup, usr_settings, merge_list):
    """
    TUPPLE2LATEXSTRING

    This function converts a tupple of openpyxl CELLs into a single row string
    of LaTeX code for inclusion in the table. It loops over each cell and appends
    the appropriate text (representing the LaTeX code) to the string which it
    returns at the end.


   Args:
        row_tup: [tupple] contains the openpyxl CELLs for a single row of the table


        merged_list = list of cells in the row that are merged together


    Returns:
        A string of the row cells formatted in the LaTeX style.

    """

    num_elements = len(row_tup)  # how many columns we have in the row

    str_out = ""  # initilise the output string

    merge_start_cols = merge_list[0]
    merge_end_cols = merge_list[1]
    merge_match_det = merge_list[2]

    colidx = 0
    multiidx = 0

    while colidx < num_elements:  # for each column/cell in the tupple for the row

        # Check to see if the column/cell is part of a multicolumn/row

        if colidx in merge_start_cols:
            # Multicolumn/row

            multcol_indx = merge_start_cols.index(colidx)
            value_string = clean_cell_str(merge_match_det[multcol_indx])

            colidx = merge_end_cols[multcol_indx]

            multiidx += 1

        else:
            # Get the main text for that cell.

            #########
            # Step 1: Get the "value_string" giving the text displayed in the cell
            #########

            if row_tup[colidx].value == None:
                # In this case, the cell is empty, so

                value_string = " "  # Cell is empty of value

            else:  # Case when the cell contains something
                # Get content of cell, and if needed, apply the d.p. rounding rule to the content.
                if usr_settings['roundtodp']:

                    if cell_is_value(str(row_tup[colidx].value)):
                        value_string = round_num_in_str(clean_cell_str(str(row_tup[colidx].value)), usr_settings['numdp'])
                    else:
                        value_string = clean_cell_str(str(row_tup[colidx].value))
                else:
                    value_string = clean_cell_str(str(row_tup[colidx].value))

            #########
            # Step 2: Apply formatting to the cell's value
            #########

            # The cell might have special formatting applied to the value inside it (e.g. bold text).
            # Apply the LaTeX version of this formatting to the string

            # Apply bold font if needed
            if row_tup[colidx].font.__dict__['b']:
                value_string = "\\textbf{" + value_string + "}"

            # Apply italicize if needed
            if row_tup[colidx].font.__dict__['i']:
                value_string = "\\textit{" + value_string + "}"

            # Apply font color
            if isinstance(row_tup[colidx].font.color.rgb, str):
                value_string = "\\textcolor[HTML]{" + row_tup[colidx].font.color.rgb[2:] + "}{" + value_string + "}"

            # Cell background color
            if isinstance(row_tup[colidx].fill.start_color.index, int):  # built in color

                # Currently cannot handle this case
                value_string = value_string

            else:

                if row_tup[colidx].fill.start_color.index is not '00000000':
                    value_string = "\\cellcolor[HTML]{" + row_tup[colidx].fill.start_color.index[2:] + "}{" + value_string + "}"

        #########
        # Step 3: Now that we have to LaTeX code for that cell/column, append it to the string for the entire row.
        #########

        # Append formatted string for this cell to the string out
        str_out += value_string

        # If this isnt the last element, add cell divider
        if colidx < num_elements - 1:
            str_out += " \t & \t "

        colidx += 1

    # Now that we have looped over all elements, add on line ending code for the end of the row string
    str_out += " \\\ \n"

    return str_out


def check_for_vline(col_tup, loc):
    """
    CHECK_FOR_VLINE

    Look for vertical lines down the entire length of the column.

    We do this by looping over all the cells, and then counting how many of them have a vertical line in location "loc"

    Args:
        row_tup: [tupple] contains the openpyxl CELLs for a single row of the table


    Returns:
        True/False
    """

    num_rows = len(col_tup)  # Number of rows in the column

    ID = 0  # Initialise count

    for rownum in range(0, num_rows):  # For each row

        # Check to see if there is a border style in location "loc"
        if col_tup[rownum].border.__dict__[loc].border_style != None:
            # Add one to our count
            ID += 1

    # Check to see if every row has a border style in location "loc"
    if ID == num_rows:
        return (True)
    else:
        return (False)

# A test of the function:
# out = check_for_vline(sheet.columns[2],'left')
# print(out)


def create_cline_code(cell_has_rule, booktabs=False):
    """
    CREATE_CLINE_CODE

    Creates the code for horizontal lines that do not span the entire length of the table.


    Args:
        cell_has_rule: [list] whose elements are True/False for each cell indicating
                            whether the horizontal rule includes this cell.

        booktabs=False: True/False. Should the code return code for the booktabs package or regular LaTeX?


    Returns:
        A string containing the code needed to draw the horizontal lines. E.g. "\cmidrule(r){1-4} \cmidrule(r){6-9} \n"
    """

    # Initialize the output string
    str_out = ''


    num_column = len(cell_has_rule)  # How many elements in the row

    # Create a flag to indicate whether we are starting a new cmidrule/cline or not.
    start_flag = 1  # we are looking for the next True

    for colind in range(0, num_column):  # For each column/cell

        if (cell_has_rule[colind] == True) & (start_flag == 1):

            # We are initializing a new cmidrule/cline

            colnum = colind + 1

            if booktabs == True:
                str_out += '\\cmidrule(r){' + str(colnum) + '-'
            else:
                str_out += '\\cline{' + str(colnum) + '-'

            start_flag = 0  # Turn off flag since now we are going to be looking for where this particular cline ends

            continue

        elif (cell_has_rule[colind] == False) & (start_flag == 1):

            # This is the case when we are searching for a new cline to begin, but there is some cells where it hasnt started yet
            str_out += ''  # do nothing

            continue

        elif cell_has_rule[colind] == False:

            # We have found a column/cell without a cline, so end the current open cline

            str_out += str(colind) + '} \t '

            start_flag = 1  # Turn flag back on so we are searching for a new cline start
            continue

        elif (cell_has_rule[colind] == True) & (colind == num_column - 1):

            # If we get to the end of the table, and the column/cell still has a cline, end the cline.
            # last one is True

            str_out += str(num_column) + '} \t '

        else:
            # Code should never reach this here.

            str_out + ' Err in create_cline_code'
            continue

    # End the line and return the string
    str_out += ' \n'

    return (str_out)


def create_horzrule_code(row_tup, loc, merge_start_cols, merge_end_cols, usr_settings, top_row=False, bottom_row=False):
    """
    CREATE_HORZRULE_CODE

    Creates the code for horizontal lines in a particular row of the table. Will return
    a string of LaTeX code that is either a horizontal line spanning the entire width of
    the table, or code to make the horizontal line(s) span only part of the table.


    Args:
        row_tup: [tuple] containing a row of the table.

        loc: [string] either 'top' or 'bottom' to indicate where (relative to this particular
                row) we should check for any horizontal lines

        row_merged_details_list: [list] details of where any merged cells start and end. Only the first cell contains the boarder info, so we copy that to all merged cells.

        usr_settings: [dictionary] user settings - tells us to use booktabs code or not.



    Returns:
        A string containing the code needed to draw the horizontal line(s) for that particular row.
    """

    '''
    Inputs:
    * row_tup : tupple containing the excel row

    * loc : String , either 'top' or 'bottom'

    * usr_settings : dirctionary, so we know if the user wants booktabs
    '''

    num_column = len(row_tup)

    #########
    # Step 1: Find which cells have horizontal rules
    #########

    # Contruct a list with True/False elements to indicate if the horizontal rule applies to that cell.

    cell_has_rule = []  # Pre-allocate list



    for colnum in range(0, num_column):

        # Check to see if column is a subsequent merged column (and hence doesnt have rule information for)
        cond_1 = [ colnum > x for x in merge_start_cols]
        cond_2 = [colnum  <= x for x in merge_end_cols]

        cond_combine = []
        for i in range(0, len(cond_1)):
            cond_combine.append(cond_1[i] is True and cond_2[i] is True)

        if any(x is True for x in cond_combine):  # cell is a subsequent merged cell
            cell_has_rule.append(cell_has_rule[-1])

        else:

            if row_tup[colnum].border.__dict__[loc].border_style != None:
                cell_has_rule.append(True)
            else:
                cell_has_rule.append(False)

    if sum(cell_has_rule) == 0:  # If there are no rules and any cell, there is no line here, so return a blank string

        return ('')

    else:  # There exists some horizontal rule on at least part of the row, so return the appropriate LaTeX code




        # If user has specified booktabs
        if usr_settings['booktabs'] == True:

            if sum(cell_has_rule) == num_column:

                return ('\midrule \n')

            else:

                return (create_cline_code(cell_has_rule, booktabs=True))






        else:

            if sum(cell_has_rule) == num_column:
                return '\hline \n'

            else:

                return create_cline_code(cell_has_rule, booktabs=False)


def get_merged_cells(sheet):
    start_row = []
    start_col = []

    end_row = []
    end_col = []

    latex_code = []

    if len(sheet.merged_cell_ranges) == 0:
        return ([[], [], [], [], []])

    for merge_ in sheet.merged_cell_ranges:  # For each merge in the sheet

        # Split the location string of the merge, and convert it it to an index number (e.g. "A3")
        merge_loc_str = re.split(':', merge_)

        merge_cell_start_str = merge_loc_str[0]
        merge_cell_end_str = merge_loc_str[1]

        # convert string to col/row index numbers
        start_coord = openpyxl.utils.coordinate_to_tuple(merge_loc_str[0])
        end_coord = openpyxl.utils.coordinate_to_tuple(merge_loc_str[1])

        start_row.append(start_coord[0] - 1)
        start_col.append(start_coord[1] - 1)

        end_row.append(end_coord[0] - 1)
        end_col.append(end_coord[1] - 1)

        value_string = sheet[merge_loc_str[0]].value

        ###HERE###

        if sheet[merge_loc_str[0]].font.__dict__['b']:
            value_string = "\\textbf{" + value_string + "}"

        # Apply italicize if needed
        if sheet[merge_loc_str[0]].font.__dict__['i']:
            value_string = "\\textit{" + value_string + "}"

        # Get span of multicolumn
        multi_col_length = end_coord[1] - start_coord[1] + 1

        # Get alignment
        halign = sheet[merge_loc_str[0]].alignment.__dict__['horizontal'][0]  # get the first letter

        latex_code.append('\multicolumn{' + str(multi_col_length) + '}{' + halign + '}{' + value_string + '}')

    return [start_row, start_col, end_row, end_col, latex_code]


def pick_col_text_alignment(col_tup):
    """
    PICK_COL_TEXT_ALIGNMENT

    For a given column, choose the alignment (left, center, right) based
    on the alignment choice of the majority of the cells


    Args:
        col_tup: [tuple] containing a column of the table.

    Returns:
        A string ('l'/'c'/'r') indicating the alignment to use
    """

    max_column = len(col_tup)

    # Preallocate counters
    count_left = 0
    count_center = 0
    count_right = 0

    # Loop over each row, and count the alignment types
    for rn in range(0, max_column):

        # If the user doesnt speicify an alignment in Excel, we see the alignment
        # choice as "None". So let us assign default values. If a number, align
        # right, if not, align left.

        if col_tup[rn].alignment.__dict__['horizontal'] is None:

            # Check to see if the value is a number
            if col_tup[rn].value is None:
                align_val = 'ignore'
            elif is_number(col_tup[rn].value):
                align_val = 'right'
            else:
                align_val = 'left'

        else:
            align_val = col_tup[rn].alignment.__dict__['horizontal']

        if align_val in ['left']:

            count_left += 1

        elif align_val in ['center']:

            count_center += 1

        elif align_val in ['right']:

            count_right += 1

    # Find the maximum, in the case of a tie, we break the tie by the order: L,C,R
    max_count = max([count_left, count_center, count_right])

    if count_left == max_count:
        return 'l'
    elif count_center == max_count:
        return 'c'
    elif count_right == max_count:
        return 'r'


def round_num_in_str(str_in, num_dp):
    """
    ROUND_NUM_IN_STR

    For a given string, round any number to the appropriate number of d.p.

    Args:
        str_in: [string] string containing numbers to round

        num_dp: [scalar] number of decimal places to round each number to

    Returns:
        A string where the numbers in str_in have been rounded.
    """

    str_out = str_in

    # Extract a list of all numbers in the string
    list_found_num_str = re.findall("\d+[\.]\d*", str_in) # Add a question mark behind the "]" to round all numbers (even those without a DP)


    list_found_scientific_num_str = re.findall("\d+[\.]*\d*e[+-]\d*", str_in)

    list_found = list_found_num_str + list_found_scientific_num_str

    # Create a list of the found numbers rounded to the appropriate d.p.
    list_nums = [float(s) for s in list_found]

    str_format = '%.' + str(num_dp) + 'f'

    for ii in range(0, len(list_found)):
        # For each number found, substitute in the rounded number
        str_out = re.sub(list_found[ii], str_format % list_nums[ii], str_out)


    return str_out


def all_nones(iterable):
    """
    Tells us if every value within the tuple iterable is None (missing)
    :param iterable: 
    :return: 
    """

    for element in iterable:
        if not element.value is None:
            return False
    return True


def get_table_dimensions(sheet):
    """
    GET_TABLE_DIMENSIONS

    The table within the sheet may not start in cell A1. This function finds the location of the table within the sheet
    by looking for the upper-left and bottom-right most cells that have content. It returns the location of these two
    corner cells.

    :param sheet: Excel worksheet object
    :return:    start_row_idx: row number of the upper-left most cell that contains something
                start_col_idx: column number of the upper-left most cell that contains something
                end_row_idx: row number of the bottom-right most cell that contains something
                end_col_idx: column number of the bottom-right most cell that contains something
    """

    # Pre-allocate starting indices
    start_col_idx = 0
    start_row_idx = 0

    # Pre-allocate end indices (adjust for python starting index at zero, and Excel starting at 1)
    end_col_idx = sheet.max_column - 1
    end_row_idx = sheet.max_row - 1

    # Trim off any empty columns at the end of the table
    for col_num in range(sheet.max_column - 1, -1, -1):

        if all_nones(list(sheet.columns)[col_num]):
            # Trim the column for the sheet
            end_col_idx = end_col_idx - 1
        else:
            # current final column has a value so stop trimming
            break

    # Trim off any empty rows at the end of the table
    for row_num in range(sheet.max_row - 1, -1, -1):

        if all_nones(list(sheet.rows)[row_num]):
            # Trim the column for the sheet
            end_row_idx = end_row_idx - 1
        else:
            # current final column has value so stop trimming
            break

    # Trim off any empty columns at the start of the table
    for col_num in range(0, sheet.max_column):

        if all_nones(list(sheet.columns)[col_num]):
            # Trim the column for the sheet
            start_col_idx = start_col_idx + 1
        else:
            # current final column has value so stop trimming
            break

    # Trim off any empty rows at the start of the table
    for row_num in range(0, sheet.max_row):

        if all_nones(list(sheet.rows)[row_num]):
            # Trim the column for the sheet
            start_row_idx = start_row_idx + 1
        else:
            # current final column has value so stop trimming
            break

    return start_row_idx, start_col_idx, end_row_idx, end_col_idx


def create_column(table_in, col_idx):
    """
    CREATE_COLUMN

    Takes a table and returns a tuple containing only a single column of the table. Useful for analysing a single
    column of the table


    :param table_in: table to extract column from
    :param col_idx: index of the column to be extracted
    :return: tuple of just column col_idx
    """

    nrows = len(table_in)  # number of rows in the table

    col = []  # preallocate

    for row_num in range(0, nrows):

        col += [table_in[row_num][col_idx]]

    return tuple(col)
